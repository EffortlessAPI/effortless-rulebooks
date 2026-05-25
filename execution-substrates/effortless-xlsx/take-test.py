#!/usr/bin/env python3
"""
effortless-xlsx test runner.

Reads the workbook produced by the LICENSED rulebook-to-xlsx transpiler at
licensed-effortless-tools/xlsx/ and emits test-answers/*.json. The cached
formula values stored in the workbook are produced by the Effortless tool
itself; if they are missing, we ask LibreOffice (if installed) to recalc.

If neither cached values nor LibreOffice are available, the substrate emits
nulls for calculated fields and scores 0 on them. That is the correct,
honest behaviour — this substrate is meant to demonstrate trust in the
licensed Effortless tool's output, not to reimplement Excel evaluation
locally.
"""

import glob
import json
import re
import shutil
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("FATAL: openpyxl not installed. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

import os

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent

# Both env vars are required — defaulting to the repo's own dirs silently
# uses a different domain's tool output and testing data.
_ERB_DOMAIN = os.environ.get("ERB_DOMAIN_DIR")
if not _ERB_DOMAIN:
    raise RuntimeError(
        "ERB_DOMAIN_DIR is not set. effortless-xlsx/take-test.py must be invoked "
        "by the orchestrator with ERB_DOMAIN_DIR pointing at the active domain's "
        "directory (e.g. rulebook-examples/acme-llc)."
    )
XLSX_TOOL_DIR = Path(_ERB_DOMAIN) / "effortless-xlsx"
if not XLSX_TOOL_DIR.exists():
    raise FileNotFoundError(
        f"effortless-xlsx tool dir not found at {XLSX_TOOL_DIR}. "
        f"Run 'effortless build' in {_ERB_DOMAIN} with rulebooktoxlsx enabled."
    )

_ERB_TESTING = os.environ.get("ERB_TESTING_DIR")
if not _ERB_TESTING:
    raise RuntimeError(
        "ERB_TESTING_DIR is not set. effortless-xlsx/take-test.py must be "
        "invoked by the orchestrator with ERB_TESTING_DIR pointing at the "
        "active domain's testing/ directory."
    )
BLANK_TESTS_DIR = Path(_ERB_TESTING) / "blank-tests"
TEST_ANSWERS_DIR = Path(_ERB_TESTING) / SCRIPT_DIR.name / "test-answers"


def to_snake_case(name: str) -> str:
    s1 = re.sub('([^_])([A-Z][a-z]+)', r'\1_\2', name)
    return re.sub('([a-z0-9])([A-Z])', r'\1_\2', s1).lower()


def find_workbook() -> Path:
    candidates = sorted(XLSX_TOOL_DIR.glob("*.xlsx"))
    if not candidates:
        print(f"FATAL: no .xlsx file in {XLSX_TOOL_DIR}", file=sys.stderr)
        print("  Run: effortless build (with rulebooktoxlsx enabled)", file=sys.stderr)
        sys.exit(1)
    return candidates[0]


def force_recalc_with_libreoffice(workbook_path: Path) -> None:
    """If LibreOffice (soffice) is installed, open + save the workbook to
    force formula recalculation. No-op otherwise."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        print("  (LibreOffice not installed — relying on cached formula values)")
        return
    print(f"  Forcing formula recalc via {soffice}...")
    subprocess.run(
        [soffice, "--headless", "--calc", "--convert-to", "xlsx",
         "--outdir", str(workbook_path.parent), str(workbook_path)],
        check=False, capture_output=True,
    )


def sheet_to_records(ws) -> list:
    """Read a worksheet into a list of {snake_case_field: value} dicts.

    Row 1 is treated as the header row. Empty rows are skipped.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [to_snake_case(str(h)) if h else None for h in rows[0]]
    records = []
    for row in rows[1:]:
        if all(v is None or v == "" for v in row):
            continue
        record = {}
        for header, value in zip(headers, row):
            if header is None:
                continue
            record[header] = value
        records.append(record)
    return records


def match_records_to_blank_tests(sheet_records: list, blank_records: list, pk_field: str) -> list:
    """Order/filter sheet_records to match the blank_tests ordering by PK.

    Ensures the test-answers JSON aligns with answer-keys JSON.
    """
    by_pk = {r.get(pk_field): r for r in sheet_records if r.get(pk_field) is not None}
    matched = []
    for blank in blank_records:
        pk = blank.get(pk_field)
        sheet_row = by_pk.get(pk)
        if sheet_row is None:
            # Workbook is missing this record entirely → record fails honestly.
            matched.append({**blank})
        else:
            merged = {**blank}
            for k, v in sheet_row.items():
                merged[k] = v
            matched.append(merged)
    return matched


def main():
    print("Effortless-XLSX Substrate Test Runner")
    print("=" * 50)
    print()

    workbook_path = find_workbook()
    print(f"Workbook: {workbook_path.relative_to(PROJECT_ROOT)}")

    force_recalc_with_libreoffice(workbook_path)

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet_names = wb.sheetnames
    print(f"Sheets: {sheet_names}")
    print()

    TEST_ANSWERS_DIR.mkdir(exist_ok=True)
    total_records = 0
    entity_count = 0

    for blank_path in sorted(BLANK_TESTS_DIR.glob("*.json")):
        entity = blank_path.stem
        with open(blank_path) as f:
            blank_records = json.load(f)
        if not blank_records:
            continue

        # Find the worksheet for this entity. Try several name variants.
        candidates = [
            entity,
            entity.title(),
            entity.capitalize(),
            ''.join(p.capitalize() for p in entity.split('_')),
        ]
        ws = None
        for name in candidates:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            # Match case-insensitively
            for sn in wb.sheetnames:
                if sn.lower() == entity.lower():
                    ws = wb[sn]
                    break
        if ws is None:
            print(f"  Skipping {entity} (no matching sheet in workbook)")
            continue

        sheet_records = sheet_to_records(ws)
        pk_field = next((k for k in blank_records[0].keys() if k.endswith("_id")), None) or list(blank_records[0].keys())[0]
        merged = match_records_to_blank_tests(sheet_records, blank_records, pk_field)

        out_path = TEST_ANSWERS_DIR / blank_path.name
        with open(out_path, "w") as f:
            json.dump(merged, f, indent=2, default=str)
        print(f"  -> {entity}: {len(merged)} records")
        total_records += len(merged)
        entity_count += 1

    print()
    print(f"effortless-xlsx: Processed {entity_count} entities, {total_records} total records")


if __name__ == "__main__":
    main()
